import os
import cv2
import face_recognition
from datetime import datetime
from openpyxl import Workbook, load_workbook

# Function to create a folder for storing student images
def create_student_folder(student_name):
    folder_path = f"student_images/{student_name}"
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
    return folder_path

# Function to capture and save student face images after 6 seconds
def get_student_face_information(student_name):
    folder_path = create_student_folder(student_name)
    camera = cv2.VideoCapture(0)

    print("Get ready to capture image in 6 seconds...")
    cv2.waitKey(6000)  # Wait for 6 seconds before capturing

    ret, frame = camera.read()
    if not ret:
        print("Error: Unable to capture image.")
        return

    image_path = os.path.join(folder_path, f"{student_name}.jpg")
    cv2.imwrite(image_path, frame)
    print(f"Image captured for {student_name}")

    camera.release()
    cv2.destroyAllWindows()

# Function to mark attendance in Excel sheet
def mark_attendance(wb, student_name):
    now = datetime.now()
    current_date = now.strftime("%Y-%m-%d")
    current_time = now.strftime("%H:%M:%S")

    ws = wb.active

    # Check if attendance for the current date and student already exists
    for row in ws.iter_rows(values_only=True):
        if row[0] == current_date and row[2] == student_name:
            print(f"Attendance already marked for {student_name} on {current_date}")
            return  # Skip marking attendance if already exists

    ws.append([current_date, current_time, student_name])

# Function to recognize student faces and mark attendance
def recognize_and_mark_attendance(wb):
    known_faces = []
    known_names = []

    for folder_name in os.listdir("student_images"):
        folder_path = os.path.join("student_images", folder_name)
        if os.path.isdir(folder_path):
            for image_name in os.listdir(folder_path):
                image_path = os.path.join(folder_path, image_name)
                image = cv2.imread(image_path)
                encoding = face_recognition.face_encodings(image)[0]
                known_faces.append(encoding)
                known_names.append(folder_name)

    camera = cv2.VideoCapture(0)
    start_time = datetime.now()

    while (datetime.now() - start_time).total_seconds() < 6:  # Capture attendance for 6 seconds
        ret, frame = camera.read()
        if not ret:
            continue

        face_locations = face_recognition.face_locations(frame)
        face_encodings = face_recognition.face_encodings(frame, face_locations)

        for face_encoding in face_encodings:
            matches = face_recognition.compare_faces(known_faces, face_encoding)
            name = "Unknown"

            if True in matches:
                first_match_index = matches.index(True)
                name = known_names[first_match_index]

            top, right, bottom, left = face_locations[0]
            cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)
            cv2.putText(frame, name, (left, top - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 0), 2)

            mark_attendance(wb, name)

        cv2.imshow("Attendance System", frame)
        cv2.waitKey(1)

    camera.release()
    cv2.destroyAllWindows()

if __name__ == "__main__":
    # Check if attendance Excel file exists
    if os.path.exists("attendance.xlsx"):
        wb = load_workbook("attendance.xlsx")
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Date", "Time", "Student Name"])
        wb.save("attendance.xlsx")

    print("Welcome to Facial Recognition Based Attendance System")
    while True:
        print("\nMenu:")
        print("1. Get Student Face Information")
        print("2. Take Attendance")
        print("3. Exit")
        choice = input("Enter your choice: ")

        if choice == '1':
            student_name = input("Enter student name: ")
            get_student_face_information(student_name)
        elif choice == '2':
            recognize_and_mark_attendance(wb)
            wb.save("attendance.xlsx")  # Save the workbook after marking attendance
        elif choice == '3':
            print("Exiting...")
            break
        else:
            print("Invalid choice. Please try again.")
